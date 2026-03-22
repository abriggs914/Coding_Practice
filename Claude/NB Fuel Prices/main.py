import streamlit as st
import pandas as pd
import plotly.graph_objects as go
import openpyxl
from datetime import datetime, timedelta
import numpy as np

# ── Page Config ──────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="NB Petroleum Prices",
    page_icon="⛽",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown("""
<style>
[data-testid="stMetricValue"] { font-size: 1.5rem; }
.metric-card { background: #1e1e2e; border-radius: 10px; padding: 12px 16px; margin: 4px 0; }
</style>
""", unsafe_allow_html=True)

# ── Load & Parse Data ─────────────────────────────────────────────────────────
XLSX_PATH = "Historical_Petroleum_Prices.xlsx"

FUEL_LABELS = {
    "Regular Unleaded Benchmark Price (New York) / Prix repères pour l'essence ordinaire (New York)": "Regular - Benchmark (NY)",
    "Regular Unleaded  Wholesale Price / Prix de gros pour l'essence ordinaire": "Regular - Wholesale",
    "Regular Unleaded  Maximum price / Prix maximum pour l'essence ordinaire": "Regular - Maximum",
    "Regular Unleaded  Maximum with Delivery / Prix maximum avec livraison pour l'essence ordianire": "Regular Self Serve ⭐",
    "Mid-grade  Benchmark Price (New York) / Prix repères pour l'essence intermédiaire (New York)": "Mid-grade - Benchmark (NY)",
    "Mid-grade  Wholesale Price / Prix de gros pour l'essence intermédiaire ": "Mid-grade - Wholesale",
    "Mid-grade Maximum price / Prix maximum pour l'essence intermédiaire": "Mid-grade - Maximum",
    "Mid-grade Maximum with Delivery / Prix maximum avec livraison pour l'essence intermédiaire": "Mid-grade Self Serve",
    "Premium Benchmark Price (New York) / Prix repères pour l'essence supercarburant (New York)": "Premium - Benchmark (NY)",
    "Premium  Wholesale Price / Prix de gros pour l'essence supercarburant": "Premium - Wholesale",
    "Premium  Maximum price / Prix maximum pour l'essence supercarburant": "Premium - Maximum",
    "Premium Maximum with Delivery / Prix maximum avec livraison pour l'essence supercarburant": "Premium Self Serve",
    "Ultra-Low Sulphur Diesel  Benchmark Price (New York) / Prix repères pour le carburant diesel à trés faible teneur en soufre (New York)": "Diesel - Benchmark (NY)",
    "Ultra-Low Sulphur Diesel Wholesale Price / Prix de gros pour le carburant diesel à trés faible teneur en soufre ": "Diesel - Wholesale",
    "Ultra-Low Sulphur Diesel Maximum Price / Prix maximum pour le carburant diesel à trés faible teneur en soufre ": "Diesel - Maximum",
    "Ultra-Low Sulphur Diesel Maximum with Delivery / Prix maximum avec livraison pour le carburant diesel à trés faible teneur en soufre ": "Diesel Self Serve",
    "Furnace Oil  Benchmark Price (New York) / Prix repères pour le mazout (New York)": "Furnace Oil - Benchmark (NY)",
    "Furnace Oil Wholesale Price / Prix de gros pour le mazout ": "Furnace Oil - Wholesale",
    "Furnace Oil  Maximum price / Prix maximum pour le mazout ": "Furnace Oil - Maximum",
    "Furnace OilMaximum with Delivery / Prix maximum avec livraison pour le mazout ": "Furnace Oil w/ Delivery",
    "Propane  Benchmark Price (Sarnia, Ont.) / Prix repères pour le propane (Sarnia, Ont.)": "Propane - Benchmark (Sarnia)",
    "Propane  Wholesale Price / Prix de gros pour le propane": "Propane - Wholesale",
    "Propane Maximum price / Prix maximum pour le propane": "Propane - Maximum",
    "Propane Maximum with Delivery / Prix maximum avec livraison pour le propane": "Propane w/ Delivery",
}

REGULAR_SELF_SERVE_KEY = "Regular Unleaded  Maximum with Delivery / Prix maximum avec livraison pour l'essence ordianire"
REGULAR_SELF_SERVE_LABEL = "Regular Self Serve ⭐"

@st.cache_data
def load_all_data():
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    sheets = wb.sheetnames  # ['Current', '2023', '2022', ...]

    all_records = {}  # label -> {date: value}

    for sheet_name in sheets:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(min_row=3, values_only=True))
        if len(rows) < 3:
            continue

        # Row index 0 = row 3 in sheet = dates
        date_row = rows[0]
        # Columns B onward (index 1+)
        dates = []
        for v in date_row[1:]:
            if v is None:
                dates.append(None)
            elif isinstance(v, datetime):
                # Filter out obviously corrupt dates (e.g. 2105 typo in source data)
                if v.year > 2030:
                    dates.append(None)
                else:
                    dates.append(v.date())
            else:
                dates.append(None)

        # Data starts at row index 2 (row 5 in sheet)
        for data_row in rows[2:]:
            label = data_row[0]
            if label is None or not isinstance(label, str):
                continue
            label = label.strip()
            if label not in all_records:
                all_records[label] = {}
            for i, val in enumerate(data_row[1:]):
                if i < len(dates) and dates[i] is not None and val is not None:
                    try:
                        all_records[label][dates[i]] = float(val)
                    except (TypeError, ValueError):
                        pass

    wb.close()

    # Convert to DataFrames keyed by display label
    dfs = {}
    for raw_label, date_dict in all_records.items():
        display_label = FUEL_LABELS.get(raw_label, raw_label)
        if not date_dict:
            continue
        s = pd.Series(date_dict, name="price")
        s.index = pd.to_datetime(list(s.index))
        s = s.sort_index()
        dfs[display_label] = s

    return dfs

# ── Load Data ─────────────────────────────────────────────────────────────────
with st.spinner("Loading petroleum price data..."):
    data = load_all_data()

all_labels = sorted(data.keys())
regular_label = REGULAR_SELF_SERVE_LABEL

# ── Header ────────────────────────────────────────────────────────────────────
st.title("⛽ New Brunswick Petroleum Prices")
st.caption("Source: New Brunswick Energy & Utilities Board (nbeub.ca) · Prices in ¢/L")

# ── Sidebar ───────────────────────────────────────────────────────────────────
st.sidebar.header("📊 Chart Settings")

selected_fuels = st.sidebar.multiselect(
    "Fuel Grade(s) to Plot",
    options=all_labels,
    default=[regular_label],
    help="Select one or more fuel grades to display on the chart."
)

min_date = min(s.index.min() for s in data.values())
max_date = max(s.index.max() for s in data.values())

date_range = st.sidebar.date_input(
    "Date Range",
    value=(min_date.to_pydatetime().date(), max_date.to_pydatetime().date()),
    min_value=min_date.to_pydatetime().date(),
    max_value=max_date.to_pydatetime().date(),
)

show_markers = st.sidebar.checkbox("Show Data Points", value=False)
show_grid = st.sidebar.checkbox("Show Grid", value=True)

st.sidebar.divider()
st.sidebar.header("🔍 Price Change Analysis")
window_days = st.sidebar.slider(
    "Analysis Window (days)",
    min_value=5,
    max_value=365,
    value=30,
    step=5,
    help="Analyse the biggest price swings over this many days."
)

analysis_fuel = st.sidebar.selectbox(
    "Fuel Grade for Analysis",
    options=all_labels,
    index=all_labels.index(regular_label) if regular_label in all_labels else 0,
)

# ── Metrics Row ───────────────────────────────────────────────────────────────
st.subheader("📈 Current Metrics – Regular Self Serve")

if regular_label in data:
    reg = data[regular_label]
    latest_val = reg.iloc[-1]
    latest_date = reg.index[-1]

    # week ago, month ago, year ago
    def val_ago(days):
        target = latest_date - pd.Timedelta(days=days)
        subset = reg[reg.index <= target]
        return subset.iloc[-1] if len(subset) > 0 else None

    week_ago = val_ago(7)
    month_ago = val_ago(30)
    year_ago = val_ago(365)

    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("Current Price", f"{latest_val:.1f} ¢/L",
                  help=f"As of {latest_date.date()}")
    with col2:
        delta = f"{latest_val - week_ago:+.1f} ¢" if week_ago is not None else "N/A"
        st.metric("vs. 1 Week Ago", f"{week_ago:.1f} ¢/L" if week_ago else "—", delta=delta if week_ago else None)
    with col3:
        delta = f"{latest_val - month_ago:+.1f} ¢" if month_ago is not None else "N/A"
        st.metric("vs. 1 Month Ago", f"{month_ago:.1f} ¢/L" if month_ago else "—", delta=delta if month_ago else None)
    with col4:
        delta = f"{latest_val - year_ago:+.1f} ¢" if year_ago is not None else "N/A"
        st.metric("vs. 1 Year Ago", f"{year_ago:.1f} ¢/L" if year_ago else "—", delta=delta if year_ago else None)

    # All-time stats
    st.markdown("")
    c1, c2, c3, c4 = st.columns(4)
    with c1:
        st.metric("All-Time High", f"{reg.max():.1f} ¢/L",
                  help=f"{reg.idxmax().date()}")
    with c2:
        st.metric("All-Time Low", f"{reg.min():.1f} ¢/L",
                  help=f"{reg.idxmin().date()}")
    with c3:
        st.metric("Historical Average", f"{reg.mean():.1f} ¢/L")
    with c4:
        # 52-week range
        yr_data = reg[reg.index >= latest_date - pd.Timedelta(days=365)]
        st.metric("52-Week Range", f"{yr_data.min():.1f} – {yr_data.max():.1f} ¢/L")

st.divider()

# ── Main Chart ─────────────────────────────────────────────────────────────────
st.subheader("📉 Price History Chart")

if not selected_fuels:
    st.info("👈 Select at least one fuel grade from the sidebar.")
else:
    try:
        start_dt = pd.Timestamp(date_range[0])
        end_dt = pd.Timestamp(date_range[1])
    except Exception:
        start_dt = min_date
        end_dt = max_date

    colors = [
        "#4fc3f7", "#ff7043", "#66bb6a", "#ffa726", "#ab47bc",
        "#26c6da", "#ef5350", "#d4e157", "#ec407a", "#42a5f5",
    ]

    fig = go.Figure()
    for i, fuel in enumerate(selected_fuels):
        if fuel not in data:
            continue
        s = data[fuel]
        mask = (s.index >= start_dt) & (s.index <= end_dt)
        s_filtered = s[mask]
        if s_filtered.empty:
            continue
        mode = "lines+markers" if show_markers else "lines"
        fig.add_trace(go.Scatter(
            x=s_filtered.index,
            y=s_filtered.values,
            mode=mode,
            name=fuel,
            line=dict(color=colors[i % len(colors)], width=2),
            marker=dict(size=4),
            hovertemplate="%{x|%b %d, %Y}<br><b>%{y:.1f} ¢/L</b><extra>" + fuel + "</extra>",
        ))

    fig.update_layout(
        template="plotly_dark",
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(20,20,35,0.8)",
        xaxis=dict(
            title="Date",
            showgrid=show_grid,
            gridcolor="rgba(255,255,255,0.08)",
            tickformat="%b %Y",
        ),
        yaxis=dict(
            title="Price (¢/L)",
            showgrid=show_grid,
            gridcolor="rgba(255,255,255,0.08)",
        ),
        hovermode="x unified",
        legend=dict(
            orientation="h",
            yanchor="bottom",
            y=1.02,
            xanchor="left",
            x=0,
            bgcolor="rgba(0,0,0,0.4)",
            bordercolor="rgba(255,255,255,0.2)",
            borderwidth=1,
        ),
        margin=dict(l=60, r=20, t=40, b=60),
        height=500,
    )
    st.plotly_chart(fig, use_container_width=True)

st.divider()

# ── Price Change Analysis ──────────────────────────────────────────────────────
st.subheader(f"🔎 Biggest Price Swings – {analysis_fuel} over {window_days} days")

if analysis_fuel in data:
    s = data[analysis_fuel].dropna()

    results = []
    dates_arr = s.index.to_numpy()
    vals_arr = s.values

    for i in range(len(dates_arr)):
        t0 = dates_arr[i]
        cutoff = t0 + pd.Timedelta(days=window_days)
        # find closest date within window
        future_mask = (dates_arr > t0) & (dates_arr <= cutoff)
        if not future_mask.any():
            continue
        future_idx = np.where(future_mask)[0]
        # biggest positive and negative change within window
        changes = vals_arr[future_idx] - vals_arr[i]
        max_change_idx = np.argmax(changes)
        min_change_idx = np.argmin(changes)
        results.append({
            "start_date": t0,
            "end_date": dates_arr[future_idx[max_change_idx]],
            "start_price": vals_arr[i],
            "end_price": vals_arr[future_idx[max_change_idx]],
            "change": changes[max_change_idx],
        })
        results.append({
            "start_date": t0,
            "end_date": dates_arr[future_idx[min_change_idx]],
            "start_price": vals_arr[i],
            "end_price": vals_arr[future_idx[min_change_idx]],
            "change": changes[min_change_idx],
        })

    if results:
        df_results = pd.DataFrame(results)
        df_results["abs_change"] = df_results["change"].abs()

        top_increase = df_results[df_results["change"] > 0].nlargest(1, "change").iloc[0]
        top_decrease = df_results[df_results["change"] < 0].nsmallest(1, "change").iloc[0]

        col_a, col_b = st.columns(2)

        with col_a:
            st.markdown("### 📈 Biggest Increase")
            inc_days = (top_increase["end_date"] - top_increase["start_date"]).days
            st.success(
                f"**+{top_increase['change']:.1f} ¢/L** over {inc_days} days\n\n"
                f"From **{top_increase['start_price']:.1f} ¢/L** on "
                f"{pd.Timestamp(top_increase['start_date']).strftime('%b %d, %Y')}\n\n"
                f"To **{top_increase['end_price']:.1f} ¢/L** on "
                f"{pd.Timestamp(top_increase['end_date']).strftime('%b %d, %Y')}"
            )

        with col_b:
            st.markdown("### 📉 Biggest Decrease")
            dec_days = (top_decrease["end_date"] - top_decrease["start_date"]).days
            st.error(
                f"**{top_decrease['change']:.1f} ¢/L** over {dec_days} days\n\n"
                f"From **{top_decrease['start_price']:.1f} ¢/L** on "
                f"{pd.Timestamp(top_decrease['start_date']).strftime('%b %d, %Y')}\n\n"
                f"To **{top_decrease['end_price']:.1f} ¢/L** on "
                f"{pd.Timestamp(top_decrease['end_date']).strftime('%b %d, %Y')}"
            )

        # Show top 10 swings table
        st.markdown("#### Top 10 Largest Swings (any direction)")
        top10 = df_results.nlargest(10, "abs_change")[
            ["start_date", "end_date", "start_price", "end_price", "change"]
        ].copy()
        top10["start_date"] = pd.to_datetime(top10["start_date"]).dt.strftime("%b %d, %Y")
        top10["end_date"] = pd.to_datetime(top10["end_date"]).dt.strftime("%b %d, %Y")
        top10["change"] = top10["change"].apply(lambda x: f"{x:+.1f} ¢/L")
        top10["start_price"] = top10["start_price"].apply(lambda x: f"{x:.1f} ¢/L")
        top10["end_price"] = top10["end_price"].apply(lambda x: f"{x:.1f} ¢/L")
        top10.columns = ["Start Date", "End Date", "Start Price", "End Price", "Change"]
        top10 = top10.reset_index(drop=True)
        top10.index += 1
        st.dataframe(top10, use_container_width=True)

st.divider()
st.caption("Data sourced from the New Brunswick Energy & Utilities Board (nbeub.ca). Prices in Canadian cents per litre (¢/L).")